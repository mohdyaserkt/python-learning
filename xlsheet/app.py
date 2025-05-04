import openpyxl as xl
wb=xl.load_workbook('transactions (1).xlsx')
sheet=wb["Sheet1"]
cell=sheet.cell(1,1)
print(cell.value)

for row in range(1,sheet.max_row+1):
    newcell=sheet.cell(row,4)
    if row==1:
        newcell.value="newPrice"
        continue

    cell=sheet.cell(row,3)
    newPrice=cell.value*0.9
    newcell.value=newPrice
    print(newcell.value)
wb.save("new.xlsx")
